import os
import re
from openpyxl import load_workbook
import logging

# Project-wide import strategy: Always use absolute imports from the project root (e.g., from src.utils import ...)
# This ensures imports work whether scripts are run directly or as modules.
from src.utils import setup_logging, get_logger, update_token_counters, get_token_usage, print_token_usage, reset_token_counters, setup_project_directory, generate_timestamp
setup_project_directory()

# Set up logging
logger = get_logger("Stable_Excel_Parsing")

def sanitize(cell):
    """Convert cell to string, strip whitespace, and replace special characters."""
    if cell is None:
        return ""
    return re.sub(r'[\n\r\t]', ' ', str(cell).strip())

def excel_to_markdown_table(file_path, sheet_name=None):
    wb = load_workbook(file_path, data_only=True)
    markdown_tables = []

    # Determine the sheets to process
    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets

    for sheet in sheets:
        raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not raw_rows:
            table = "No data found."
        else:
            # Determine the maximum number of columns
            max_cols = max(len(row) for row in raw_rows)

            # Sanitize and pad all rows
            rows = [
                [sanitize(cell) for cell in row] + [""] * (max_cols - len(row))
                for row in raw_rows
            ]

            # Calculate consistent column widths
            col_widths = [max(len(row[i]) for row in rows) for i in range(max_cols)]

            # Format Markdown table
            header = "| " + " | ".join(f"{rows[0][i]:<{col_widths[i]}}" for i in range(max_cols)) + " |"
            separator = "|-" + "-|-".join("-" * col_widths[i] for i in range(max_cols)) + "-|"
            table_lines = [header, separator]

            for row in rows[1:]:
                line = "| " + " | ".join(f"{row[i]:<{col_widths[i]}}" for i in range(max_cols)) + " |"
                table_lines.append(line)

            table = "\n".join(table_lines)

        markdown_tables.append(f"Sheet: {sheet.title}\n\n{table}\n")

    return "\n".join(markdown_tables)


# === Example usage ===
path = ""  # Replace with your Excel file
markdown_output = excel_to_markdown_table(path)
output_file = "src/_02_parsing/helpdesk_export_openpyxl.md"

with open(output_file, "w", encoding="utf-8") as md_file:
    md_file.write(markdown_output)

logger.info(f"Markdown file '{output_file}' has been created.", extra={"icon": "✅"})
